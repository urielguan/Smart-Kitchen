#!/usr/bin/env python3
import hashlib
import json
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import pymysql
from openpyxl import load_workbook


JSON_SOURCE_VERSION = "json_data_vision_251206_Qwen2-5-VL-72B-Instruct"
EXCEL_SOURCE_VERSION = "china-food-composition-standard-6th-edition-excel"
JSON_BASE_DIR = Path(__file__).resolve().parents[3] / "doc" / "china-food-composition-data-main" / JSON_SOURCE_VERSION
EXCEL_BASE_DIR = Path(__file__).resolve().parents[3] / "doc" / "中国食物成分表  标准版  第6版"
NAME_NORMALIZE_RE = re.compile(r"[\s\W_（）()【】\[\]《》]+", re.UNICODE)
CANDIDATE_LEVEL1_NAME = "候选补充食品"
CANDIDATE_LEVEL2_NAME = "01食品营养成分数据库"

DB_CONFIG = {
    "host": "172.31.25.155",
    "port": 3306,
    "user": "root",
    "password": "YingziOS#2026",
    "database": "smart_food_safety",
    "charset": "utf8mb4",
    "autocommit": False,
}


CREATE_TABLE_SQL = [
    """
    CREATE TABLE IF NOT EXISTS food_category (
      id BIGINT NOT NULL AUTO_INCREMENT COMMENT '分类ID',
      parent_id BIGINT DEFAULT NULL COMMENT '父分类ID',
      category_code VARCHAR(64) NOT NULL COMMENT '分类编码',
      category_name VARCHAR(100) NOT NULL COMMENT '分类名称',
      category_level TINYINT NOT NULL COMMENT '分类层级：1/2',
      sort_order INT NOT NULL DEFAULT 0 COMMENT '排序',
      source_file VARCHAR(255) DEFAULT NULL COMMENT '来源文件',
      status VARCHAR(20) NOT NULL DEFAULT 'active' COMMENT '状态',
      created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
      updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
      deleted TINYINT NOT NULL DEFAULT 0 COMMENT '逻辑删除',
      PRIMARY KEY (id),
      UNIQUE KEY uk_food_category_code (category_code),
      KEY idx_food_category_parent (parent_id),
      KEY idx_food_category_level (category_level)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci COMMENT='标准食品分类表'
    """,
    """
    CREATE TABLE IF NOT EXISTS food_item (
      id BIGINT NOT NULL AUTO_INCREMENT COMMENT '标准食品ID',
      food_code VARCHAR(32) NOT NULL COMMENT '标准食品编码',
      food_name VARCHAR(150) NOT NULL COMMENT '标准食品名称',
      category_level1_id BIGINT DEFAULT NULL COMMENT '一级分类ID',
      category_level2_id BIGINT DEFAULT NULL COMMENT '二级分类ID',
      edible_ratio DECIMAL(5,2) DEFAULT NULL COMMENT '可食部比例',
      energy_kcal DECIMAL(10,2) DEFAULT NULL COMMENT '热量（千卡/100g）',
      protein DECIMAL(10,2) DEFAULT NULL COMMENT '蛋白质（g/100g）',
      fat DECIMAL(10,2) DEFAULT NULL COMMENT '脂肪（g/100g）',
      carbohydrate DECIMAL(10,2) DEFAULT NULL COMMENT '碳水化合物（g/100g）',
      dietary_fiber DECIMAL(10,2) DEFAULT NULL COMMENT '膳食纤维（g/100g）',
      sodium DECIMAL(10,2) DEFAULT NULL COMMENT '钠（mg/100g）',
      vitamin_a DECIMAL(10,2) DEFAULT NULL COMMENT '维生素A（ug/100g）',
      vitamin_b1 DECIMAL(10,2) DEFAULT NULL COMMENT '维生素B1（mg/100g）',
      vitamin_b2 DECIMAL(10,2) DEFAULT NULL COMMENT '维生素B2（mg/100g）',
      vitamin_c DECIMAL(10,2) DEFAULT NULL COMMENT '维生素C（mg/100g）',
      vitamin_e DECIMAL(10,2) DEFAULT NULL COMMENT '维生素E（mg/100g）',
      calcium DECIMAL(10,2) DEFAULT NULL COMMENT '钙（mg/100g）',
      iron DECIMAL(10,2) DEFAULT NULL COMMENT '铁（mg/100g）',
      zinc DECIMAL(10,2) DEFAULT NULL COMMENT '锌（mg/100g）',
      source_file VARCHAR(255) DEFAULT NULL COMMENT '来源文件',
      source_version VARCHAR(100) DEFAULT NULL COMMENT '来源版本',
      raw_payload JSON DEFAULT NULL COMMENT '原始JSON快照',
      status VARCHAR(20) NOT NULL DEFAULT 'active' COMMENT '状态',
      created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
      updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
      deleted TINYINT NOT NULL DEFAULT 0 COMMENT '逻辑删除',
      PRIMARY KEY (id),
      UNIQUE KEY uk_food_item_code (food_code),
      KEY idx_food_item_level1 (category_level1_id),
      KEY idx_food_item_level2 (category_level2_id)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci COMMENT='标准食品主表'
    """,
    """
    CREATE TABLE IF NOT EXISTS material_food_mapping (
      id BIGINT NOT NULL AUTO_INCREMENT COMMENT '映射ID',
      material_id BIGINT NOT NULL COMMENT '业务物料ID',
      food_item_id BIGINT NOT NULL COMMENT '标准食品ID',
      match_status VARCHAR(20) NOT NULL DEFAULT 'confirmed' COMMENT '映射状态',
      confirmed_by BIGINT DEFAULT NULL COMMENT '确认人',
      confirmed_at DATETIME DEFAULT NULL COMMENT '确认时间',
      remark VARCHAR(300) DEFAULT NULL COMMENT '备注',
      created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
      updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
      deleted TINYINT NOT NULL DEFAULT 0 COMMENT '逻辑删除',
      PRIMARY KEY (id),
      UNIQUE KEY uk_material_food_mapping (material_id),
      KEY idx_material_food_food_item (food_item_id)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci COMMENT='业务物料-标准食品映射表'
    """,
    """
    CREATE TABLE IF NOT EXISTS recipe_nutrition_result (
      id BIGINT NOT NULL AUTO_INCREMENT COMMENT '结果ID',
      recipe_id BIGINT NOT NULL COMMENT '菜谱ID',
      calc_version INT NOT NULL DEFAULT 1 COMMENT '计算版本',
      calories DECIMAL(10,2) DEFAULT NULL COMMENT '热量',
      protein DECIMAL(10,2) DEFAULT NULL COMMENT '蛋白质',
      carbohydrate DECIMAL(10,2) DEFAULT NULL COMMENT '碳水',
      fat DECIMAL(10,2) DEFAULT NULL COMMENT '脂肪',
      sodium DECIMAL(10,2) DEFAULT NULL COMMENT '钠',
      fiber DECIMAL(10,2) DEFAULT NULL COMMENT '膳食纤维',
      vitamin_a DECIMAL(10,2) DEFAULT NULL COMMENT '维生素A',
      vitamin_b1 DECIMAL(10,2) DEFAULT NULL COMMENT '维生素B1',
      vitamin_b2 DECIMAL(10,2) DEFAULT NULL COMMENT '维生素B2',
      vitamin_c DECIMAL(10,2) DEFAULT NULL COMMENT '维生素C',
      vitamin_d DECIMAL(10,2) DEFAULT NULL COMMENT '维生素D',
      vitamin_e DECIMAL(10,2) DEFAULT NULL COMMENT '维生素E',
      calcium DECIMAL(10,2) DEFAULT NULL COMMENT '钙',
      iron DECIMAL(10,2) DEFAULT NULL COMMENT '铁',
      zinc DECIMAL(10,2) DEFAULT NULL COMMENT '锌',
      nutrition_score INT DEFAULT NULL COMMENT '营养评分',
      pass_status VARCHAR(20) DEFAULT NULL COMMENT '达标状态',
      data_completeness DECIMAL(5,2) DEFAULT NULL COMMENT '数据完整度',
      missing_material_count INT NOT NULL DEFAULT 0 COMMENT '缺失物料数',
      missing_materials TEXT DEFAULT NULL COMMENT '缺失物料列表',
      calculated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '计算时间',
      created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
      updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
      deleted TINYINT NOT NULL DEFAULT 0 COMMENT '逻辑删除',
      PRIMARY KEY (id),
      UNIQUE KEY uk_recipe_nutrition_result_recipe (recipe_id)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci COMMENT='菜谱营养计算结果表'
    """,
]

WMS_MATERIAL_COLUMNS = {
    "food_item_id": "ALTER TABLE wms_material ADD COLUMN food_item_id BIGINT DEFAULT NULL COMMENT '关联标准食品ID' AFTER material_category",
    "calories": "ALTER TABLE wms_material ADD COLUMN calories DECIMAL(10,2) DEFAULT NULL COMMENT '热量（千卡/100g）' AFTER remark",
    "protein": "ALTER TABLE wms_material ADD COLUMN protein DECIMAL(10,2) DEFAULT NULL COMMENT '蛋白质（g/100g）' AFTER calories",
    "carbohydrate": "ALTER TABLE wms_material ADD COLUMN carbohydrate DECIMAL(10,2) DEFAULT NULL COMMENT '碳水化合物（g/100g）' AFTER protein",
    "fat": "ALTER TABLE wms_material ADD COLUMN fat DECIMAL(10,2) DEFAULT NULL COMMENT '脂肪（g/100g）' AFTER carbohydrate",
    "sodium": "ALTER TABLE wms_material ADD COLUMN sodium DECIMAL(10,2) DEFAULT NULL COMMENT '钠（mg/100g）' AFTER fat",
    "fiber": "ALTER TABLE wms_material ADD COLUMN fiber DECIMAL(10,2) DEFAULT NULL COMMENT '膳食纤维（g/100g）' AFTER sodium",
    "vitamin_a": "ALTER TABLE wms_material ADD COLUMN vitamin_a DECIMAL(10,2) DEFAULT NULL COMMENT '维生素A（ug/100g）' AFTER fiber",
    "vitamin_b1": "ALTER TABLE wms_material ADD COLUMN vitamin_b1 DECIMAL(10,2) DEFAULT NULL COMMENT '维生素B1（mg/100g）' AFTER vitamin_a",
    "vitamin_b2": "ALTER TABLE wms_material ADD COLUMN vitamin_b2 DECIMAL(10,2) DEFAULT NULL COMMENT '维生素B2（mg/100g）' AFTER vitamin_b1",
    "vitamin_c": "ALTER TABLE wms_material ADD COLUMN vitamin_c DECIMAL(10,2) DEFAULT NULL COMMENT '维生素C（mg/100g）' AFTER vitamin_b2",
    "vitamin_e": "ALTER TABLE wms_material ADD COLUMN vitamin_e DECIMAL(10,2) DEFAULT NULL COMMENT '维生素E（mg/100g）' AFTER vitamin_c",
    "calcium": "ALTER TABLE wms_material ADD COLUMN calcium DECIMAL(10,2) DEFAULT NULL COMMENT '钙（mg/100g）' AFTER vitamin_e",
    "iron": "ALTER TABLE wms_material ADD COLUMN iron DECIMAL(10,2) DEFAULT NULL COMMENT '铁（mg/100g）' AFTER calcium",
    "zinc": "ALTER TABLE wms_material ADD COLUMN zinc DECIMAL(10,2) DEFAULT NULL COMMENT '锌（mg/100g）' AFTER iron",
    "nutrition_source_type": "ALTER TABLE wms_material ADD COLUMN nutrition_source_type VARCHAR(30) DEFAULT NULL COMMENT '营养来源类型' AFTER zinc",
    "nutrition_source_ref_id": "ALTER TABLE wms_material ADD COLUMN nutrition_source_ref_id BIGINT DEFAULT NULL COMMENT '营养来源引用ID' AFTER nutrition_source_type",
    "nutrition_synced_at": "ALTER TABLE wms_material ADD COLUMN nutrition_synced_at DATETIME DEFAULT NULL COMMENT '营养同步时间' AFTER nutrition_source_ref_id",
}


def md5_code(parent_id, level, name):
    raw = f"{'ROOT' if parent_id is None else parent_id}:{level}:{name}"
    return "FC_" + hashlib.md5(raw.encode("utf-8")).hexdigest()[:16]


def decimal_value(raw):
    if raw is None:
        return None
    value = str(raw).strip()
    if not value or value == "—":
        return None
    if value.lower() == "tr":
        return Decimal("0")
    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return None


def normalize_food_name(name):
    if not name:
        return ""
    return NAME_NORMALIZE_RE.sub("", str(name).lower())


def build_candidate_food_code(name):
    return "C01" + hashlib.md5(normalize_food_name(name).encode("utf-8")).hexdigest()[:8].upper()


def append_source_value(existing, new_value):
    if not new_value:
        return existing
    values = []
    if existing:
        values.extend([item.strip() for item in str(existing).split(";") if item.strip()])
    if new_value not in values:
        values.append(new_value)
    return "; ".join(values)


def merge_raw_payload(existing_raw, supplement_payload, source_type):
    supplement_payload = dict(supplement_payload)
    supplement_payload["sourceType"] = source_type
    if not existing_raw:
        return json_dumps(supplement_payload)
    try:
        existing = json.loads(existing_raw)
    except json.JSONDecodeError:
        existing = {"baseRaw": existing_raw}
    if isinstance(existing, dict) and "supplements" in existing:
        payload = existing
    else:
        payload = {"base": existing, "supplements": []}
    payload.setdefault("supplements", []).append(supplement_payload)
    return json_dumps(payload)


def json_dumps(payload):
    return json.dumps(payload, ensure_ascii=False, default=json_default)


def json_default(value):
    if isinstance(value, Decimal):
        return float(value)
    raise TypeError(f"Object of type {value.__class__.__name__} is not JSON serializable")


def ensure_schema(cur):
    for sql in CREATE_TABLE_SQL:
        cur.execute(sql)

    for column_name, alter_sql in WMS_MATERIAL_COLUMNS.items():
        cur.execute(
            """
            SELECT COUNT(*)
            FROM information_schema.columns
            WHERE table_schema = %s AND table_name = 'wms_material' AND column_name = %s
            """,
            (DB_CONFIG["database"], column_name),
        )
        if cur.fetchone()[0] == 0:
            cur.execute(alter_sql)


def load_item_cache(cur):
    cur.execute(
        """
        SELECT id, food_code, food_name, category_level1_id, category_level2_id, edible_ratio, energy_kcal,
               protein, fat, carbohydrate, dietary_fiber, sodium, vitamin_a, vitamin_b1, vitamin_b2,
               vitamin_c, vitamin_e, calcium, iron, zinc, source_file, source_version, raw_payload
          FROM food_item
         WHERE deleted = 0
        """
    )
    items = []
    by_code = {}
    by_name = {}
    for row in cur.fetchall():
        item = {
            "id": row[0],
            "food_code": row[1],
            "food_name": row[2],
            "category_level1_id": row[3],
            "category_level2_id": row[4],
            "edible_ratio": row[5],
            "energy_kcal": row[6],
            "protein": row[7],
            "fat": row[8],
            "carbohydrate": row[9],
            "dietary_fiber": row[10],
            "sodium": row[11],
            "vitamin_a": row[12],
            "vitamin_b1": row[13],
            "vitamin_b2": row[14],
            "vitamin_c": row[15],
            "vitamin_e": row[16],
            "calcium": row[17],
            "iron": row[18],
            "zinc": row[19],
            "source_file": row[20],
            "source_version": row[21],
            "raw_payload": row[22],
        }
        items.append(item)
        by_code[item["food_code"]] = item
        if item["food_name"]:
            by_name.setdefault(normalize_food_name(item["food_name"]), item)
    return items, by_code, by_name


def load_category_cache(cur):
    cur.execute("SELECT id, category_code, category_name, category_level, source_file FROM food_category WHERE deleted = 0")
    cache = {}
    for row in cur.fetchall():
        cache[row[1]] = {
            "id": row[0],
            "category_code": row[1],
            "category_name": row[2],
            "category_level": row[3],
            "source_file": row[4],
        }
    return cache


def upsert_category(cur, cache, parent_id, name, level, sort_order, file_name):
    code = md5_code(parent_id, level, name)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cached = cache.get(code)
    if cached:
        cur.execute(
            """
            UPDATE food_category
               SET source_file=%s, updated_at=%s
             WHERE id=%s
            """,
            (append_source_value(cached.get("source_file"), file_name), now, cached["id"]),
        )
        cached["source_file"] = append_source_value(cached.get("source_file"), file_name)
        return cached["id"]

    cur.execute(
        """
        INSERT INTO food_category
            (parent_id, category_code, category_name, category_level, sort_order, source_file, status, created_at, updated_at, deleted)
        VALUES
            (%s, %s, %s, %s, %s, %s, 'active', %s, %s, 0)
        """,
        (parent_id, code, name, level, sort_order, file_name, now, now),
    )
    category_id = cur.lastrowid
    cache[code] = {
        "id": category_id,
        "category_code": code,
        "category_name": name,
        "category_level": level,
        "source_file": file_name,
    }
    return category_id


def update_item_cache(item, by_code, by_name):
    if item.get("food_code"):
        by_code[item["food_code"]] = item
    if item.get("food_name"):
        by_name.setdefault(normalize_food_name(item["food_name"]), item)


def insert_food_item(cur, item):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        """
        INSERT INTO food_item
            (food_code, food_name, category_level1_id, category_level2_id, edible_ratio, energy_kcal,
             protein, fat, carbohydrate, dietary_fiber, sodium, vitamin_a, vitamin_b1, vitamin_b2,
             vitamin_c, vitamin_e, calcium, iron, zinc, source_file, source_version, raw_payload,
             status, created_at, updated_at, deleted)
        VALUES
            (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'active', %s, %s, 0)
        """,
        (
            item["food_code"], item["food_name"], item["category_level1_id"], item["category_level2_id"],
            item["edible_ratio"], item["energy_kcal"], item["protein"], item["fat"], item["carbohydrate"],
            item["dietary_fiber"], item["sodium"], item["vitamin_a"], item["vitamin_b1"], item["vitamin_b2"],
            item["vitamin_c"], item["vitamin_e"], item["calcium"], item["iron"], item["zinc"], item["source_file"],
            item["source_version"], item["raw_payload"], now, now,
        ),
    )
    item["id"] = cur.lastrowid


def update_food_item(cur, item):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        """
        UPDATE food_item
           SET food_name=%s, category_level1_id=%s, category_level2_id=%s, edible_ratio=%s,
               energy_kcal=%s, protein=%s, fat=%s, carbohydrate=%s, dietary_fiber=%s, sodium=%s,
               vitamin_a=%s, vitamin_b1=%s, vitamin_b2=%s, vitamin_c=%s, vitamin_e=%s,
               calcium=%s, iron=%s, zinc=%s, source_file=%s, source_version=%s, raw_payload=%s,
               updated_at=%s
         WHERE id=%s
        """,
        (
            item["food_name"], item["category_level1_id"], item["category_level2_id"], item["edible_ratio"],
            item["energy_kcal"], item["protein"], item["fat"], item["carbohydrate"], item["dietary_fiber"],
            item["sodium"], item["vitamin_a"], item["vitamin_b1"], item["vitamin_b2"], item["vitamin_c"],
            item["vitamin_e"], item["calcium"], item["iron"], item["zinc"], item["source_file"],
            item["source_version"], item["raw_payload"], now, item["id"],
        ),
    )


def fill_missing_fields(item, payload):
    filled = 0
    field_map = [
        ("edible_ratio", "edible_ratio"),
        ("energy_kcal", "energy_kcal"),
        ("protein", "protein"),
        ("fat", "fat"),
        ("carbohydrate", "carbohydrate"),
        ("dietary_fiber", "dietary_fiber"),
        ("sodium", "sodium"),
        ("vitamin_a", "vitamin_a"),
        ("vitamin_b1", "vitamin_b1"),
        ("vitamin_b2", "vitamin_b2"),
        ("vitamin_c", "vitamin_c"),
        ("vitamin_e", "vitamin_e"),
        ("calcium", "calcium"),
        ("iron", "iron"),
        ("zinc", "zinc"),
    ]
    for item_key, payload_key in field_map:
        if item.get(item_key) is None and payload.get(payload_key) is not None:
            item[item_key] = payload[payload_key]
            filled += 1
    return filled


def import_json_baseline(cur, category_cache, item_by_code, item_by_name, stats):
    files = sorted(JSON_BASE_DIR.glob("*.json"))
    stats["file_count"] += len(files)
    cur.execute("SELECT COUNT(*) FROM food_item WHERE deleted = 0 AND source_version = %s", (JSON_SOURCE_VERSION,))
    if cur.fetchone()[0] > 0:
        return

    top_sort = 10
    for file in files:
        file_name = file.stem
        category_raw = file_name.replace("merged_", "")
        parts = category_raw.split("-", 1)
        if len(parts) < 2:
            stats["skipped_count"] += 1
            continue
        level1_id = upsert_category(cur, category_cache, None, parts[0], 1, top_sort, file.name)
        level2_id = upsert_category(cur, category_cache, level1_id, parts[1], 2, top_sort, file.name)
        top_sort += 10
        rows = json.loads(file.read_text(encoding="utf-8"))
        stats["source_stats"][file.name] = len(rows)
        for row in rows:
            food_code = str(row.get("foodCode") or "").strip()
            if not food_code:
                stats["skipped_count"] += 1
                continue
            if food_code in item_by_code:
                stats["duplicate_count"] += 1
                continue
            item = {
                "food_code": food_code,
                "food_name": str(row.get("foodName") or "").strip() or None,
                "category_level1_id": level1_id,
                "category_level2_id": level2_id,
                "edible_ratio": decimal_value(row.get("edible")),
                "energy_kcal": decimal_value(row.get("energyKCal")),
                "protein": decimal_value(row.get("protein")),
                "fat": decimal_value(row.get("fat")),
                "carbohydrate": decimal_value(row.get("CHO")),
                "dietary_fiber": decimal_value(row.get("dietaryFiber")),
                "sodium": decimal_value(row.get("Na")),
                "vitamin_a": decimal_value(row.get("vitaminA")),
                "vitamin_b1": decimal_value(row.get("thiamin")),
                "vitamin_b2": decimal_value(row.get("riboflavin")),
                "vitamin_c": decimal_value(row.get("vitaminC")),
                "vitamin_e": decimal_value(row.get("vitaminETotal")),
                "calcium": decimal_value(row.get("Ca")),
                "iron": decimal_value(row.get("Fe")),
                "zinc": decimal_value(row.get("Zn")),
                "source_file": file.name,
                "source_version": JSON_SOURCE_VERSION,
                "raw_payload": json_dumps(row),
            }
            insert_food_item(cur, item)
            update_item_cache(item, item_by_code, item_by_name)
            stats["new_item_count"] += 1


def column_labels(ws):
    labels = []
    row_zh = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row_en = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    max_len = max(len(row_zh), len(row_en))
    for idx in range(max_len):
        zh = "" if idx >= len(row_zh) or row_zh[idx] is None else str(row_zh[idx]).strip()
        en = "" if idx >= len(row_en) or row_en[idx] is None else str(row_en[idx]).strip()
        labels.append(f"{zh}|{en}".lower().replace(" ", ""))
    return labels


def find_col(labels, *candidates):
    normalized = [candidate.lower().replace(" ", "") for candidate in candidates]
    for idx, label in enumerate(labels):
        if any(candidate in label for candidate in normalized):
            return idx
    return -1


def cell_value(row_values, idx):
    if idx < 0 or idx >= len(row_values):
        return None
    value = row_values[idx]
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if value is None:
        return None
    return str(value).strip()


def excel_payload_from_row(row_values, labels, file_name, sheet_name, level1_name, level2_name):
    energy_kcal_raw = cell_value(row_values, find_col(labels, "能量|energy"))
    payload = {
        "food_code": cell_value(row_values, find_col(labels, "食物编码", "foodcode")),
        "food_name": cell_value(row_values, find_col(labels, "食物名称", "foodname")),
        "edible_ratio": decimal_value(cell_value(row_values, find_col(labels, "食部", "edible"))),
        "energy_kcal": decimal_value(energy_kcal_raw),
        "protein": decimal_value(cell_value(row_values, find_col(labels, "蛋白质", "protein"))),
        "fat": decimal_value(cell_value(row_values, find_col(labels, "脂肪", "fat"))),
        "carbohydrate": decimal_value(cell_value(row_values, find_col(labels, "碳水化合物", "cho"))),
        "dietary_fiber": decimal_value(cell_value(row_values, find_col(labels, "膳食纤维", "dietaryfiber"))),
        "sodium": decimal_value(cell_value(row_values, find_col(labels, "钠", "|na"))),
        "vitamin_a": decimal_value(cell_value(row_values, find_col(labels, "总维生素a", "vitamina"))),
        "vitamin_b1": decimal_value(cell_value(row_values, find_col(labels, "硫胺素", "thiamin"))),
        "vitamin_b2": decimal_value(cell_value(row_values, find_col(labels, "核黄素", "riboflavin"))),
        "vitamin_c": decimal_value(cell_value(row_values, find_col(labels, "维生素c", "vitaminc"))),
        "vitamin_e": decimal_value(cell_value(row_values, find_col(labels, "维生素e", "total"))),
        "calcium": decimal_value(cell_value(row_values, find_col(labels, "钙", "|ca"))),
        "iron": decimal_value(cell_value(row_values, find_col(labels, "铁", "|fe"))),
        "zinc": decimal_value(cell_value(row_values, find_col(labels, "锌", "|zn"))),
        "file_name": file_name,
        "sheet_name": sheet_name,
        "level1_name": level1_name,
        "level2_name": level2_name,
    }
    return payload


def import_excel_standard(cur, category_cache, item_by_code, item_by_name, stats):
    files = sorted(path for path in EXCEL_BASE_DIR.glob("*.xlsx") if not path.name.startswith("01"))
    stats["file_count"] += len(files)
    sort_order = 1000
    for file in files:
        processed = 0
        workbook = load_workbook(file, read_only=True, data_only=True)
        try:
            for ws in workbook.worksheets:
                labels = column_labels(ws)
                level1_name = cell_value(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), 0)
                if not level1_name:
                    continue
                level1_id = upsert_category(cur, category_cache, None, level1_name, 1, sort_order, file.name)
                current_level2 = None
                level2_sort = sort_order
                for row_values in ws.iter_rows(min_row=5, values_only=True):
                    first = cell_value(row_values, 0)
                    second = cell_value(row_values, 1)
                    if not first and not second:
                        continue
                    if first and not second:
                        current_level2 = first
                        level2_sort += 10
                        continue
                    if not current_level2:
                        stats["skipped_count"] += 1
                        continue
                    level2_id = upsert_category(cur, category_cache, level1_id, current_level2, 2, level2_sort, file.name)
                    payload = excel_payload_from_row(row_values, labels, file.name, ws.title, level1_name, current_level2)
                    food_code = payload["food_code"]
                    food_name = payload["food_name"]
                    if not food_code or not food_name:
                        stats["skipped_count"] += 1
                        continue
                    processed += 1
                    item = item_by_code.get(food_code) or item_by_name.get(normalize_food_name(food_name))
                    if item is None:
                        item = {
                            **payload,
                            "food_code": food_code,
                            "food_name": food_name,
                            "category_level1_id": level1_id,
                            "category_level2_id": level2_id,
                            "source_file": file.name,
                            "source_version": EXCEL_SOURCE_VERSION,
                            "raw_payload": json_dumps(payload),
                        }
                        insert_food_item(cur, item)
                        update_item_cache(item, item_by_code, item_by_name)
                        stats["new_item_count"] += 1
                        continue

                    filled = fill_missing_fields(item, payload)
                    if not item.get("category_level1_id"):
                        item["category_level1_id"] = level1_id
                    if not item.get("category_level2_id"):
                        item["category_level2_id"] = level2_id
                    if filled > 0:
                        item["source_file"] = append_source_value(item.get("source_file"), file.name)
                        item["source_version"] = append_source_value(item.get("source_version"), EXCEL_SOURCE_VERSION)
                        item["raw_payload"] = merge_raw_payload(item.get("raw_payload"), payload, "excel_supplement")
                        update_food_item(cur, item)
                        stats["supplemented_field_count"] += filled
                    else:
                        stats["duplicate_count"] += 1
        finally:
            workbook.close()
        stats["source_stats"][file.name] = processed
        sort_order += 100


def import_candidate_excel(cur, item_by_name, stats):
    file = EXCEL_BASE_DIR / "01食品营养成分数据库.xlsx"
    if not file.exists():
        return
    stats["file_count"] += 1
    workbook = load_workbook(file, read_only=True, data_only=True)
    processed = 0
    try:
        ws = workbook.worksheets[0]
        category_cache = load_category_cache(cur)
        level1_id = upsert_category(cur, category_cache, None, CANDIDATE_LEVEL1_NAME, 1, 9000, file.name)
        level2_id = upsert_category(cur, category_cache, level1_id, CANDIDATE_LEVEL2_NAME, 2, 9010, file.name)
        for row_values in ws.iter_rows(min_row=2, values_only=True):
            food_name = cell_value(row_values, 1)
            if not food_name:
                continue
            processed += 1
            item = item_by_name.get(normalize_food_name(food_name))
            payload = {
                "edible_ratio": decimal_value(cell_value(row_values, 8)),
                "energy_kcal": (decimal_value(cell_value(row_values, 2)) / Decimal("4.184")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                if decimal_value(cell_value(row_values, 2)) is not None else None,
                "protein": decimal_value(cell_value(row_values, 3)),
                "fat": decimal_value(cell_value(row_values, 4)),
                "carbohydrate": decimal_value(cell_value(row_values, 5)),
                "sodium": decimal_value(cell_value(row_values, 6)),
                "file_name": file.name,
                "sheet_name": ws.title,
                "level1_name": "候选补充",
                "level2_name": "第6版整理库",
            }
            if item is None:
                item = {
                    "food_code": build_candidate_food_code(food_name),
                    "food_name": food_name,
                    "category_level1_id": level1_id,
                    "category_level2_id": level2_id,
                    "edible_ratio": payload["edible_ratio"],
                    "energy_kcal": payload["energy_kcal"],
                    "protein": payload["protein"],
                    "fat": payload["fat"],
                    "carbohydrate": payload["carbohydrate"],
                    "dietary_fiber": None,
                    "sodium": payload["sodium"],
                    "vitamin_a": None,
                    "vitamin_b1": None,
                    "vitamin_b2": None,
                    "vitamin_c": None,
                    "vitamin_e": None,
                    "calcium": None,
                    "iron": None,
                    "zinc": None,
                    "source_file": file.name,
                    "source_version": EXCEL_SOURCE_VERSION,
                    "raw_payload": json_dumps({**payload, "food_name": food_name, "sourceType": "candidate_excel_insert"}),
                }
                insert_food_item(cur, item)
                item_by_name[normalize_food_name(food_name)] = item
                stats["new_item_count"] += 1
                continue
            filled = fill_missing_fields(item, payload)
            if filled > 0:
                item["source_file"] = append_source_value(item.get("source_file"), file.name)
                item["source_version"] = append_source_value(item.get("source_version"), EXCEL_SOURCE_VERSION)
                item["raw_payload"] = merge_raw_payload(item.get("raw_payload"), payload, "candidate_excel")
                update_food_item(cur, item)
                stats["supplemented_field_count"] += filled
            else:
                stats["duplicate_count"] += 1
    finally:
        workbook.close()
    stats["source_stats"][file.name] = processed


def main():
    if not JSON_BASE_DIR.is_dir():
        raise SystemExit(f"JSON 目录不存在: {JSON_BASE_DIR}")
    if not EXCEL_BASE_DIR.is_dir():
        raise SystemExit(f"Excel 目录不存在: {EXCEL_BASE_DIR}")

    conn = pymysql.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            ensure_schema(cur)
            _, item_by_code, item_by_name = load_item_cache(cur)
            category_cache = load_category_cache(cur)
            stats = {
                "file_count": 0,
                "new_item_count": 0,
                "supplemented_field_count": 0,
                "skipped_count": 0,
                "duplicate_count": 0,
                "exception_count": 0,
                "source_stats": {},
            }

            import_json_baseline(cur, category_cache, item_by_code, item_by_name, stats)
            import_excel_standard(cur, category_cache, item_by_code, item_by_name, stats)
            import_candidate_excel(cur, item_by_name, stats)
            conn.commit()

            cur.execute("SELECT COUNT(*) FROM food_category WHERE deleted = 0")
            category_count = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM food_item WHERE deleted = 0")
            item_count = cur.fetchone()[0]

        print(
            json.dumps(
                {
                    "status": "ok",
                    "file_count": stats["file_count"],
                    "food_category_count": category_count,
                    "food_item_count": item_count,
                    "new_food_item_count": stats["new_item_count"],
                    "supplemented_field_count": stats["supplemented_field_count"],
                    "skipped_item_count": stats["skipped_count"],
                    "duplicate_item_count": stats["duplicate_count"],
                    "exception_count": stats["exception_count"],
                    "source_version": f"{JSON_SOURCE_VERSION} + {EXCEL_SOURCE_VERSION}",
                    "source_stats": stats["source_stats"],
                },
                ensure_ascii=False,
                indent=2,
                default=json_default,
            )
        )
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
